from openpyxl import load_workbook
import datetime
import re
import requests
curr_time = datetime.datetime.now()
name = "{}{}".format(curr_time.hour,curr_time.minute)
'''

from time import sleep
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>


# 微信:
# 可以使用6—20个字母、数字、下划线和减号
# 必须以字母开头（不区分大小写）
# 不支持设置中文，仅能设置1次。
def 微信(content,booksheet):
    res = re.search(r"[a-zA-Z][\w-]{5,19}", content, re.A)
    if res:
        # 14列写入此内容有6+位微信号
        # booksheet.cell(i, 1).value = "6+"
        res2 = re.search(r"([a-zA-Z][\w-]{5,19})", content, re.A)
        # res2 = re.search(r"[❤️❤][^0-9a-zA-Z\-]{0,2}([a-zA-Z][\w-]{5,19})", content, re.A)
        if res2:
            # print(content)
            value = res2[1]
            # print(value)
            booksheet.cell(i, 15).value = value


# 手机号12-13列
def phone(content,booksheet):
    # 类型一: 只要有11+位数字
    res = re.search(r"1\d{10}", content)
    if res:
        # 12列写入此内容有11+位数字
        booksheet.cell(i, 20).value = "11位"

        res2 = re.search(r"(1\d{10})\b|(1\d{10})\s|(1\d{10})\Z|(1\d{10})[^0-9]", content, re.A)
        if res2:
            for ii in res2.groups():
                if ii != None:
                    value = str(ii)

            # print(content)
            # 13列写入数据电话号码
            # print(value)
            booksheet.cell(i, 21).value = value


    # 类型二: 没有11位数字/但是有137-123...这种类型
    elif (re.search(r'－|-', content)):
        # 把-去除掉
        content = re.sub(r'－|-', "", content)
        res2 = re.search(r"(1\d{10})\b|(1\d{10})\s|(1\d{10})\Z|(1\d{10})[^0-9]", content, re.A)
        if res2:
            for ii in res2.groups():
                if ii != None:
                    value = str(ii)

            print(content)
            # 13列写入数据电话号码
            print(value)
            booksheet.cell(i, 13).value = value


# 没有联系方式
def no_联系方式(content,booksheet):
    # 只要有11+位数字
    res = re.search('[a-zA-Z 0-9_-]{6}', content)  # 没有任何数字
    if res:
        pass
    else:
        # print(res[1])
        booksheet.cell(i, 16).value = 0

# 主函数>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
workbook = load_workbook('info.xlsx')
booksheet = workbook.active #获取当前活跃的sheet,默认是第一个sheet
# 读取行数
row_max = booksheet.max_row

for i in range(1,row_max+1):

    # 获取简介内容
    content = booksheet.cell(row=i, column=10).value

    # 微信
    # res = re.search(r"❤️[^0-9a-zA-Z\-]{0,2}([0-9a-zA-Z\-]{5,}[0-9a-zA-Z\-])", content)
    try:
        # 去除空行
        content = 去除空行_特殊符号(content, booksheet)

        # 号码: 并写入12-13列
        phone(content, booksheet)

    # 微信号: 并写入14-15列
    #     微信(content, booksheet)

    # 没有任何联系方式16列

        # no_联系方式(content, booksheet)
    except:
        pass

'''
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
import pymysql


# 格式化内容
def 去除空行_特殊符号(content):
    # －-
    # content: 简介内容
    # 去除空行
    content = re.sub(r"\s+", "", content)

    # 数字方框
    content = re.sub(r"⃣️", "", content)
    content = re.sub(r'️', "", content)

    # 数字方框(第二种类型)
    content = re.sub(r"⃣", "", content)

    content = re.sub(r"一", "1", content)
    content = re.sub(r"二", "2", content)
    content = re.sub(r"三", "3", content)
    content = re.sub(r"四", "4", content)
    content = re.sub(r"五", "5", content)
    content = re.sub(r"六", "6", content)
    content = re.sub(r"七", "7", content)
    content = re.sub(r"八", "8", content)
    content = re.sub(r"九", "9", content)
    content = re.sub(r"零", "0", content)

    content = re.sub(r"①", "1", content)
    content = re.sub(r"②", "2", content)
    content = re.sub(r"③", "3", content)
    content = re.sub(r"③", "4", content)
    content = re.sub(r"⑤", "5", content)
    content = re.sub(r"⑥", "6", content)
    content = re.sub(r"⑦", "7", content)
    content = re.sub(r"⑧", "8", content)
    content = re.sub(r"⑨", "9", content)
    content = re.sub(r"⑩", "10", content)

    content = re.sub(r"❶", "1", content)
    content = re.sub(r"❷", "2", content)
    content = re.sub(r"❸", "3", content)
    content = re.sub(r"❹", "4", content)
    content = re.sub(r"❺", "5", content)
    content = re.sub(r"❻", "6", content)
    content = re.sub(r"❼", "7", content)
    content = re.sub(r"❽", "8", content)
    content = re.sub(r"❾", "9", content)
    content = re.sub(r"❿", "10", content)

    content = re.sub(r"⑪", "11", content)
    content = re.sub(r"⑫", "12", content)
    content = re.sub(r"⑬", "13", content)
    content = re.sub(r"⑭", "14", content)
    content = re.sub(r"⑮", "15", content)
    content = re.sub(r"⑯", "16", content)
    content = re.sub(r"⑰", "17", content)
    content = re.sub(r"⑱", "18", content)
    content = re.sub(r"⑲", "19", content)
    content = re.sub(r"⑳", "20", content)

    return content


# 提取联系方式
# 手机号12-13列
def get_phone(content):
    # 类型一: 只要有11+位数字
    res = re.search(r"1\d{10}", content)
    if res:

        res2 = re.search(r"(1\d{10})\b|(1\d{10})\s|(1\d{10})\Z|(1\d{10})[^0-9]", content, re.A)
        if res2:
            for ii in res2.groups():
                if ii != None:
                    phone = str(ii)

            return phone


    # 类型二: 没有11位数字/但是有137-123...这种类型
    elif (re.search(r'－|-', content)):
        # 把-去除掉
        content = re.sub(r'－|-', "", content)
        res2 = re.search(r"(1\d{10})\b|(1\d{10})\s|(1\d{10})\Z|(1\d{10})[^0-9]", content, re.A)
        if res2:
            for ii in res2.groups():
                if ii != None:
                    phone = str(ii)

            print(content)
            # 13列写入数据电话号码
            return phone


for i in range(1,1000):
    # 链接数据库
    db = pymysql.connect(host="rm-bp1278x3bc1a6ujve1o.mysql.rds.aliyuncs.com", user="test_0527",passwd="test_0527",db="test_0527")
    # 使用cursor()方法获取操作游标
    cursor = db.cursor()
    try:
        # 执行sql语句
        cursor.execute("select `id`,`ks_id`,`ks_name`,`ks_intro`,`is_auth` from ks_user where is_deal is Null limit 1")
        # 获取所有记录
        results = cursor.fetchall()

        ks_id = results[0][1]
        ks_name = results[0][2]
        ks_intro = results[0][3]
        is_auth = results[0][4]

        # 格式化内容
        ks_intro = 去除空行_特殊符号(ks_intro)

        # 正则匹配联系方式
        mobile = get_phone(ks_intro)

        if mobile == None:
            pass
        else:
            cursor.execute("UPDATE ks_user SET mobile = {} WHERE  id = {}".format(mobile,results[0][0]))
            # 上传到数据库
            payload = 'ks_id={}&ks_name={}&ks_intro={}&mobile={}&is_auth={}'.format(ks_id, ks_name, ks_intro, mobile,
                                                                                    is_auth)
            headers = {'Content-Type': 'application/x-www-form-urlencoded'}
            response = requests.post("http://112.124.127.143:8049/api/Ks/PostKsUserOne", headers=headers,
                                     data=payload.encode("utf8"))
            print(ks_name, mobile)
            print(response.text)
    except:
        pass

    cursor.execute("UPDATE ks_user SET is_deal = 1 WHERE  id = {}".format(results[0][0]))
    # 打印数据
    db.commit()
    cursor.close()

    # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>






# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>







